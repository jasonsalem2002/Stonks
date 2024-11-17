import yfinance as yf
from datetime import datetime, timedelta
import pytz
import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image as ExcelImage
import matplotlib.pyplot as plt
import mplfinance as mpf
from textblob import TextBlob
import requests
from bs4 import BeautifulSoup
import numpy as np
import os
import logging

# Setup Logging
logging.basicConfig(
    filename='stock_analysis.log',
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

# Suppress specific warnings
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", message="NotOpenSSLWarning")

# Constants
TIMEZONE = pytz.timezone('America/New_York')
TODAY = datetime.now(TIMEZONE)
ONE_YEAR_AGO = TODAY - timedelta(days=365)

# Function to fetch sentiment using TextBlob (simple implementation)
def fetch_sentiment(symbol):
    try:
        url = f'https://finviz.com/quote.ashx?t={symbol}'
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        news_table = soup.find(id='news-table')
        if not news_table:
            return "N/A"
        sentiments = []
        for row in news_table.findAll('tr')[:20]:  # Analyze latest 20 news
            text = row.a.get_text()
            blob = TextBlob(text)
            sentiments.append(blob.sentiment.polarity)
        if sentiments:
            avg_sentiment = np.mean(sentiments)
            return round(avg_sentiment, 2)
        else:
            return "N/A"
    except Exception as e:
        logging.error(f"Error fetching sentiment for {symbol}: {e}")
        return "N/A"

# Function to calculate risk metrics
def calculate_risk_metrics(hist):
    try:
        returns = hist['Close'].pct_change().dropna()
        volatility = returns.std() * np.sqrt(252)  # Annualized volatility
        cumulative_returns = (1 + returns).cumprod()
        peak = cumulative_returns.cummax()
        drawdown = (cumulative_returns - peak) / peak
        max_drawdown = drawdown.min()
        sharpe_ratio = (returns.mean() / returns.std()) * np.sqrt(252) if returns.std() != 0 else "N/A"
        return round(volatility, 4), round(max_drawdown, 4), round(sharpe_ratio, 4) if sharpe_ratio != "N/A" else "N/A"
    except Exception as e:
        logging.error(f"Error calculating risk metrics: {e}")
        return "N/A", "N/A", "N/A"

# Function to perform DCF Valuation (simple implementation)
def dcf_valuation(ticker):
    try:
        fcfs = ticker.cashflow.loc['Free Cash Flow']
        if fcfs.empty:
            return "N/A"
        recent_fcf = fcfs.iloc[0]
        growth_rate = 0.05  # Assumed growth rate
        discount_rate = 0.10  # Assumed discount rate
        terminal_value = recent_fcf * (1 + growth_rate) / (discount_rate - growth_rate)
        intrinsic_value = terminal_value / (1 + discount_rate) ** 1  # Simplified for 1 year
        return round(intrinsic_value, 2)
    except Exception as e:
        logging.error(f"Error performing DCF valuation: {e}")
        return "N/A"

# Function to get user input for stock symbols
def get_user_symbols():
    user_input = input("Enter stock symbols separated by commas (e.g., AAPL, MSFT, GOOGL): ")
    symbols = [symbol.strip().upper() for symbol in user_input.split(',') if symbol.strip()]
    if not symbols:
        print("No valid symbols entered. Exiting.")
        logging.warning("No valid symbols entered by the user.")
        exit()
    return symbols

# Function to add DataFrame to sheet with styling
def add_df_to_sheet(wb, df, sheet_name, conditional_format_cols=None):
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # Apply header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    # Apply conditional formatting
    if conditional_format_cols:
        for row in ws.iter_rows(min_row=2, min_col=conditional_format_cols[0], 
                                max_col=conditional_format_cols[1], max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    header = ws.cell(row=1, column=cell.column).value
                    if any(keyword in header for keyword in ['Change', 'Ratio', 'Score']):
                        if cell.value < 0:
                            cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")  # Red for negative
                        elif cell.value > 0:
                            cell.fill = PatternFill(start_color="CCFFCC", fill_type="solid")  # Green for positive

# Function to create individual sheets for each symbol with daily, weekly, and monthly changes
def add_individual_symbol_sheets(wb, symbol, hist):
    try:
        sheet_name = f"{symbol} Changes"
        sheet_name = sheet_name[:31]  # Excel sheet name max length is 31
        ws = wb.create_sheet(sheet_name)
        
        # Calculate Daily, Weekly, Monthly Changes
        daily_changes = hist['Close'].pct_change().dropna() * 100
        weekly_changes = hist['Close'].pct_change(periods=5).dropna() * 100
        monthly_changes = hist['Close'].pct_change(periods=21).dropna() * 100  # Approx. 21 trading days in a month
        
        # Find the maximum number of periods to align the data
        # Start from the date where all changes are available
        min_length = min(len(daily_changes), len(weekly_changes), len(monthly_changes))
        if min_length <= 0:
            logging.warning(f"Insufficient data to create changes sheet for {symbol}.")
            return
        
        # Trim the series to the minimum length
        daily_changes_trimmed = daily_changes[-min_length:]
        weekly_changes_trimmed = weekly_changes[-min_length:]
        monthly_changes_trimmed = monthly_changes[-min_length:]
        dates_trimmed = daily_changes_trimmed.index
        
        # Create DataFrame
        changes_df = pd.DataFrame({
            'Date': dates_trimmed.strftime('%Y-%m-%d'),
            'Daily Change (%)': daily_changes_trimmed.values,
            'Weekly Change (%)': weekly_changes_trimmed.values,
            'Monthly Change (%)': monthly_changes_trimmed.values
        })
        
        # Append to sheet
        for r in dataframe_to_rows(changes_df, index=False, header=True):
            ws.append(r)
        
        # Apply header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Apply conditional formatting
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=4, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")  # Red
                    elif cell.value > 0:
                        cell.fill = PatternFill(start_color="CCFFCC", fill_type="solid")  # Green
    except Exception as e:
        logging.error(f"Error creating individual sheets for {symbol}: {e}")

# Function to generate and save candlestick chart for a symbol
def generate_candlestick_chart(symbol, hist):
    try:
        mc = mpf.make_marketcolors(up='g', down='r', inherit=True)
        s = mpf.make_mpf_style(marketcolors=mc)
        candle_chart = f"{symbol}_candlestick.png"
        mpf.plot(hist, type='candle', style=s, title=f"{symbol} Candlestick Chart", volume=True, savefig=candle_chart)
        return candle_chart
    except Exception as e:
        logging.error(f"Error generating candlestick chart for {symbol}: {e}")
        return None

# Function to generate and save comparative closing price chart
def generate_comparative_chart(symbols, hist_data):
    try:
        plt.figure(figsize=(12, 8))
        for symbol, hist in hist_data.items():
            if not hist.empty:
                plt.plot(hist.index, hist['Close'], label=symbol)
        plt.title("Comparative Closing Prices")
        plt.xlabel("Date")
        plt.ylabel("Price ($)")
        plt.legend()
        plt.grid(True)
        comparative_chart = "All_Stocks_Comparative.png"
        plt.savefig(comparative_chart)
        plt.close()
        return comparative_chart
    except Exception as e:
        logging.error(f"Error generating comparative chart: {e}")
        return None

# Main function to get stock information
def get_stock_info(symbols):
    all_data = []
    technical_data = []
    sentiment_data = []
    dividend_data = []
    risk_data = []
    valuation_data = []
    charts = []
    hist_data = {}
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for symbol in symbols:
        ticker = yf.Ticker(symbol)
        
        # Check if the ticker is valid
        try:
            info = ticker.info
            # A more reliable check: ensure 'marketCap' exists and is not None
            if 'marketCap' not in info or info['marketCap'] is None:
                print(f"Warning: {symbol} might be delisted or data is unavailable. Skipping.")
                logging.warning(f"{symbol} might be delisted or data is unavailable. Skipping.")
                continue
        except Exception as e:
            print(f"Error fetching data for {symbol}: {e}")
            logging.error(f"Error fetching data for {symbol}: {e}")
            continue
        
        # Company Info
        company_name = info.get('longName', 'N/A')
        sector = info.get('sector', 'N/A')
        industry = info.get('industry', 'N/A')
        market_cap = info.get('marketCap', 'N/A')
        pe_ratio = info.get('trailingPE', 'N/A')
        peg_ratio = info.get('pegRatio', 'N/A')
        eps = info.get('trailingEps', 'N/A')
        beta = info.get('beta', 'N/A')
        revenue_growth = info.get('revenueGrowth', 'N/A')
        debt_to_equity = info.get('debtToEquity', 'N/A')
        free_cash_flow = info.get('freeCashflow', 'N/A')
        dividend_yield = round(info.get('dividendYield', 0) * 100, 2) if info.get('dividendYield') else "N/A"
        payout_ratio = info.get('payoutRatio', 'N/A')
        
        # Current Price and Historical Data
        try:
            current_price = ticker.history(period='1d')['Close'].iloc[-1]
        except Exception as e:
            current_price = "N/A"
            logging.error(f"Error fetching current price for {symbol}: {e}")
        hist = ticker.history(start=ONE_YEAR_AGO, end=TODAY)
        
        if hist.empty:
            print(f"Warning: No historical data for {symbol}. Skipping further analysis for this stock.")
            logging.warning(f"No historical data for {symbol}. Skipping further analysis.")
            continue
        
        hist_data[symbol] = hist  # Store hist for comparative chart
        
        # Percentage Changes
        try:
            price_3m = hist['Close'].iloc[-63] if len(hist) >= 63 else hist['Close'].iloc[0]
            price_6m = hist['Close'].iloc[-126] if len(hist) >= 126 else hist['Close'].iloc[0]
            price_1y = hist['Close'].iloc[0]
            pct_change_3m = round(((current_price - price_3m) / price_3m) * 100, 2)
            pct_change_6m = round(((current_price - price_6m) / price_6m) * 100, 2)
            pct_change_1y = round(((current_price - price_1y) / price_1y) * 100, 2)
        except Exception as e:
            print(f"Error calculating percentage changes for {symbol}: {e}")
            logging.error(f"Error calculating percentage changes for {symbol}: {e}")
            pct_change_3m = pct_change_6m = pct_change_1y = "N/A"
        
        # Dividend History
        dividends = ticker.dividends.tail(5)
        if not dividends.empty:
            dividends = dividends.sort_index()
            dividend_dates = dividends.index.strftime('%Y-%m-%d').tolist()
            dividend_amounts = dividends.values.tolist()
            dividend_dict = {'Date': dividend_dates, 'Amount': dividend_amounts}
        else:
            dividend_dict = {'Date': ['N/A'], 'Amount': ['N/A']}
        
        # Dividend Trend (%)
        if len(dividends) >= 2:
            try:
                dividend_trend = [round(((dividend_amounts[i] - dividend_amounts[i-1]) / dividend_amounts[i-1]) * 100, 2) 
                                  for i in range(1, len(dividend_amounts))]
                dividend_trend = dividend_trend[::-1]  # Reverse to match chronological order
            except Exception as e:
                logging.error(f"Error calculating dividend trend for {symbol}: {e}")
                dividend_trend = ["N/A"]
        else:
            dividend_trend = ["N/A"]
        
        # Risk Metrics
        volatility, max_drawdown, sharpe_ratio = calculate_risk_metrics(hist)
        
        # Sentiment Analysis
        sentiment = fetch_sentiment(symbol)
        
        # Valuation Metrics
        intrinsic_value = dcf_valuation(ticker)
        fair_value = intrinsic_value  # Simplified assumption
        
        # Technical Indicators
        hist['50_MA'] = hist['Close'].rolling(window=50).mean()
        hist['200_MA'] = hist['Close'].rolling(window=200).mean()
        ma_crossover = "Golden Cross" if hist['50_MA'].iloc[-1] > hist['200_MA'].iloc[-1] else "Death Cross"
        
        technical_data.append({
            "Stock Symbol": symbol,
            "50-Day MA": round(hist['50_MA'].iloc[-1], 2) if not pd.isna(hist['50_MA'].iloc[-1]) else "N/A",
            "200-Day MA": round(hist['200_MA'].iloc[-1], 2) if not pd.isna(hist['200_MA'].iloc[-1]) else "N/A",
            "MA Crossover": ma_crossover
        })
        
        # Add to main data list
        try:
            all_data.append({
                "Stock Symbol": symbol,
                "Company Name": company_name,
                "Sector": sector,
                "Industry": industry,
                "Market Cap": f"${market_cap:,}" if isinstance(market_cap, int) else "N/A",
                "P/E Ratio": pe_ratio if pe_ratio else "N/A",
                "PEG Ratio": peg_ratio if peg_ratio else "N/A",
                "EPS": eps if eps else "N/A",
                "Beta": beta if beta else "N/A",
                "Revenue Growth (%)": round(revenue_growth * 100, 2) if revenue_growth else "N/A",
                "Debt-to-Equity": debt_to_equity if debt_to_equity else "N/A",
                "Free Cash Flow": f"${free_cash_flow:,}" if isinstance(free_cash_flow, (int, float)) else "N/A",
                "Current Price": f"${current_price:.2f}" if isinstance(current_price, (int, float)) else "N/A",
                "Price 3 Months Ago": f"${price_3m:.2f}" if isinstance(price_3m, (int, float)) else "N/A",
                "Price 6 Months Ago": f"${price_6m:.2f}" if isinstance(price_6m, (int, float)) else "N/A",
                "Price 1 Year Ago": f"${price_1y:.2f}" if isinstance(price_1y, (int, float)) else "N/A",
                "% Change (3M)": pct_change_3m,
                "% Change (6M)": pct_change_6m,
                "% Change (1Y)": pct_change_1y,
                "Dividend Yield (%)": dividend_yield,
                "Payout Ratio": round(payout_ratio, 2) if isinstance(payout_ratio, (int, float)) else "N/A",
                "Sentiment Score": sentiment,
                "Intrinsic Value": f"${intrinsic_value:.2f}" if isinstance(intrinsic_value, (int, float)) else "N/A",
                "Fair Value": f"${fair_value:.2f}" if isinstance(fair_value, (int, float)) else "N/A",
                "Volatility": volatility,
                "Max Drawdown": max_drawdown,
                "Sharpe Ratio": sharpe_ratio,
                "MA Crossover": ma_crossover
            })
        except Exception as e:
            logging.error(f"Error adding data to main list for {symbol}: {e}")
        
        # Add to sentiment data
        sentiment_data.append({
            "Stock Symbol": symbol,
            "Sentiment Score": sentiment
        })
        
        # Add to dividend data
        for date, amount in zip(dividend_dict['Date'], dividend_dict['Amount']):
            dividend_data.append({
                "Stock Symbol": symbol,
                "Dividend Date": date,
                "Dividend Amount": amount
            })
        
        # Add to risk data
        risk_data.append({
            "Stock Symbol": symbol,
            "Volatility": volatility,
            "Max Drawdown": max_drawdown,
            "Sharpe Ratio": sharpe_ratio
        })
        
        # Add to valuation data
        valuation_data.append({
            "Stock Symbol": symbol,
            "Intrinsic Value": intrinsic_value,
            "Fair Value": fair_value
        })
        
        # Generate and save candlestick chart
        candle_chart = generate_candlestick_chart(symbol, hist)
        if candle_chart:
            charts.append(candle_chart)
        
        # Create Individual Sheets for Symbol
        add_individual_symbol_sheets(wb, symbol, hist)
    
    # Create DataFrames
    df_overview = pd.DataFrame(all_data)
    df_technical = pd.DataFrame(technical_data)
    df_sentiment = pd.DataFrame(sentiment_data)
    df_dividends = pd.DataFrame(dividend_data)
    df_risk = pd.DataFrame(risk_data)
    df_valuation = pd.DataFrame(valuation_data)
    
    # Add DataFrames to Excel workbook
    add_df_to_sheet(wb, df_overview, "Stock Overview", conditional_format_cols=(13, 15))  # Columns M to O for % Changes
    add_df_to_sheet(wb, df_technical, "Technical Indicators")
    add_df_to_sheet(wb, df_sentiment, "Sentiment Analysis")
    add_df_to_sheet(wb, df_dividends, "Dividend History")
    add_df_to_sheet(wb, df_risk, "Risk Analysis")
    add_df_to_sheet(wb, df_valuation, "Valuation Metrics")
    
    # Generate and save comparative chart
    comparative_chart = generate_comparative_chart(symbols, hist_data)
    if comparative_chart:
        charts.append(comparative_chart)
    
    # Insert Charts into the workbook
    ws_overview = wb["Stock Overview"]
    row_position = df_overview.shape[0] + 5
    for chart in charts:
        if os.path.exists(chart):
            try:
                img = ExcelImage(chart)
                img.width = 640  # Adjust as needed
                img.height = 480
                ws_overview.add_image(img, f"A{row_position}")
                row_position += 30  # Adjust the row position for next image
            except Exception as e:
                print(f"Error inserting chart {chart} into Excel: {e}")
                logging.error(f"Error inserting chart {chart} into Excel: {e}")
    
    # Save the styled workbook
    excel_filename = "Enhanced_Stock_Analysis_with_Charts.xlsx"
    try:
        wb.save(excel_filename)
        print(f"Excel report generated: '{excel_filename}'")
        logging.info(f"Excel report generated: '{excel_filename}'")
    except Exception as e:
        print(f"Error saving Excel workbook: {e}")
        logging.error(f"Error saving Excel workbook: {e}")
    
    # Clean up chart images
    for chart in charts:
        if os.path.exists(chart):
            try:
                os.remove(chart)
                logging.info(f"Deleted chart image: {chart}")
            except Exception as e:
                logging.error(f"Error deleting chart image {chart}: {e}")
    
    print("Stock analysis completed. Results saved to 'Enhanced_Stock_Analysis_with_Charts.xlsx'.")
    logging.info("Stock analysis completed successfully.")

# Example usage
if __name__ == "__main__":
    # Get user input for stock symbols
    stock_symbols = get_user_symbols()
    get_stock_info(stock_symbols)
